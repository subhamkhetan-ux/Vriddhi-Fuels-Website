"""Write a Master-Paid-shaped .xlsx.

Cell types mirror the proven writer in ``pay/index.html`` exactly, so the file
drops into "Master Paid" cleanly:

  A  Date          numeric Excel serial, displayed ``dd/mm/yyyy`` (NOT text)
  B  Customer      exact canonical name
  C  Amount Paid   plain number (``#,##0`` is display grouping only)
  D  Payment Mode  bank + mode text, e.g. "HDFC NEFT"

Row 1 is a bold header, matching the pay app; data starts at row 2.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

HEADER = ["Date", "Customer", "Amount Paid", "Payment Mode"]
DATE_FMT = "dd/mm/yyyy"
AMOUNT_FMT = "#,##0"


def write_xlsx(path: str, rows: list[dict]) -> None:
    """Write ``rows`` to ``path``. Each row needs
    ``date_serial``, ``customer``, ``amount``, ``mode``."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Paid"

    bold = Font(bold=True)
    for col, name in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = bold
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=2):
        a = ws.cell(row=i, column=1, value=int(r["date_serial"]))
        a.number_format = DATE_FMT  # serial value, date display
        ws.cell(row=i, column=2, value=r["customer"])
        c = ws.cell(row=i, column=3, value=r["amount"])
        c.number_format = AMOUNT_FMT  # plain value, grouped display
        ws.cell(row=i, column=4, value=r.get("mode", ""))

    widths = {"A": 12, "B": 34, "C": 14, "D": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(path)
