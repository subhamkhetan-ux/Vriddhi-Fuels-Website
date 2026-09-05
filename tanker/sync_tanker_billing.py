#!/usr/bin/env python3
"""Materialize the Tanker Billing customer data to ``state/tanker_billing.json``.

The bill layout on the ``Billing`` sheet is fully driven by XLOOKUPs against the
``Customers`` table plus the ``RateCard`` / ``ProductRates`` helper tables. The
tanker-bill web app never opens the workbook; it reads the JSON this script
produces. Re-running keeps the app current with the master workbook (new
customers, changed PO numbers, changed prices).

Two ways to run:

    # 1) Local file (testing, or a Mac/laptop that has the workbook synced):
    python3 tanker/sync_tanker_billing.py --xlsm "Tanker Billing.xlsm"

    # 2) Pull the live copy from Google Drive (used by the scheduled Action).
    #    Auth (recommended): an OAuth refresh token for the account that owns
    #    the file — mint it with tanker/mint_drive_token.py. A service account
    #    works too. See tanker/README.md for the one-time setup.
    GDRIVE_TOKEN=/path/gdrive_token.json DRIVE_FILE_ID=<id> \
        python3 tanker/sync_tanker_billing.py --from-drive

The workbook is read with ``data_only=True`` so cached formula values are used.
That means the copy on Drive must have been saved by Excel at least once after
any change (Excel caches the computed values on save) — which your daily
iCloud->Excel workflow already does.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
import tempfile

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(REPO_ROOT, "state", "tanker_billing.json")

# Sheet / table names in the master workbook.
CUSTOMERS_SHEET = "Customers"
CUSTOMERS_TABLE = "Customers"      # A1:M32
RATECARD_TABLE = "RateCard"        # O1:P9  (Price Tier -> Rate)
PRODUCTRATES_TABLE = "ProductRates"  # R1:S3 (Product -> Rate)
FUELTYPES_TABLE = "FuelTypes"      # U1:U4 (Fuel Type dropdown)


def _clean(v):
    """Trim strings; treat Excel's blank-placeholder ' ' as empty."""
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        return s
    return v


def _num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else round(f, 4)
    except (TypeError, ValueError):
        return None


def _table_rows(ws, table):
    """Yield dict rows for an openpyxl Table, keyed by header text."""
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [
        _clean(ws.cell(min_row, c).value) for c in range(min_col, max_col + 1)
    ]
    for r in range(min_row + 1, max_row + 1):
        values = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
        if all(_clean(v) == "" for v in values):
            continue
        yield dict(zip(headers, values))


def parse_workbook(xlsm_path: str) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=False)
    if CUSTOMERS_SHEET not in wb.sheetnames:
        raise SystemExit(
            f"sheet {CUSTOMERS_SHEET!r} not found; have: {wb.sheetnames}")
    ws = wb[CUSTOMERS_SHEET]
    tables = ws.tables

    def need(name):
        if name not in tables:
            raise SystemExit(
                f"table {name!r} not found on {CUSTOMERS_SHEET!r}; "
                f"have: {list(tables)}")
        return tables[name]

    # --- Rate card: Price Tier -> Rate ---
    rate_card = {}
    for row in _table_rows(ws, need(RATECARD_TABLE)):
        tier = _clean(row.get("Price Tier"))
        rate = _num(row.get("Rate"))
        if tier and rate is not None:
            rate_card[tier] = rate

    # --- Product rates: Product -> Rate (non-HSD products) ---
    product_rates = {}
    for row in _table_rows(ws, need(PRODUCTRATES_TABLE)):
        prod = _clean(row.get("Product"))
        rate = _num(row.get("Rate"))
        if prod and rate is not None:
            product_rates[prod] = rate

    # --- Fuel-type dropdown list ---
    fuel_types = []
    for row in _table_rows(ws, need(FUELTYPES_TABLE)):
        ft = _clean(row.get("Fuel Type"))
        if ft:
            fuel_types.append(ft)

    # --- Customers ---
    customers = []
    for row in _table_rows(ws, need(CUSTOMERS_TABLE)):
        company = _clean(row.get("Company"))
        if not company:
            continue
        tier = _clean(row.get("Price Tier"))
        hsd = _num(row.get("HSD"))
        # HSD column is itself an XLOOKUP against the rate card; if the cached
        # value is missing, fall back to resolving the tier ourselves.
        if hsd is None and tier in rate_card:
            hsd = rate_card[tier]
        # Keep all three address rows and all five payment rows *unfiltered*
        # (blanks preserved). The bill grid is positional: Address 3 (row 16)
        # is always italic, payment lines sit on fixed rows — dropping blanks
        # would shift text off the cells it must line up with.
        address = [
            _clean(row.get("Address 1")),
            _clean(row.get("Address 2")),
            _clean(row.get("Address 3")),
        ]
        payment = [_clean(row.get(f"Payment {i}")) for i in range(1, 6)]
        po_no = row.get("PO No.")
        po_no = _clean(po_no)
        customers.append(
            {
                "company": company,
                "price_tier": tier,
                "hsd_rate": hsd,
                "address": [str(a) if a != "" else "" for a in address],
                "payment": [str(p) if p != "" else "" for p in payment],
                "po_label": _clean(row.get("PO Label")),
                "po_no": "" if po_no in (None, "") else str(po_no),
            }
        )

    customers.sort(key=lambda c: c["company"].lower())

    return {
        "schema": 1,
        "fuel_types": fuel_types,
        "rate_card": rate_card,
        "product_rates": product_rates,
        "customers": customers,
    }


DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


def _load_env_json(value: str):
    """Return a dict from an env value that is either a path or inline JSON."""
    if os.path.exists(value):
        with open(value, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return json.loads(value)


def drive_credentials():
    """Build read-only Drive credentials from whichever is configured.

    Preferred (reuses the same OAuth setup as the Gmail agent):
      GDRIVE_TOKEN     an OAuth *authorized-user* JSON (path or inline) minted by
                       tanker/mint_drive_token.py — a refresh token for the
                       Google account whose Drive holds the workbook. Because it
                       is your own account, no file-sharing step is needed.

    Alternative:
      GDRIVE_SA_JSON   a service-account key JSON (path or inline). The workbook
                       (or its folder) must be shared with the account's email.
    """
    token = os.environ.get("GDRIVE_TOKEN", "").strip()
    if token:
        from google.oauth2.credentials import Credentials
        return Credentials.from_authorized_user_info(
            _load_env_json(token), scopes=DRIVE_SCOPES)

    sa = os.environ.get("GDRIVE_SA_JSON", "").strip()
    if sa:
        from google.oauth2 import service_account
        return service_account.Credentials.from_service_account_info(
            _load_env_json(sa), scopes=DRIVE_SCOPES)

    raise SystemExit(
        "No Drive credentials. Set GDRIVE_TOKEN (OAuth, recommended) or "
        "GDRIVE_SA_JSON (service account). See tanker/README.md.")


def download_from_drive(dest_path: str) -> dict:
    """Download the workbook from Google Drive.

    Auth: GDRIVE_TOKEN (OAuth, recommended) or GDRIVE_SA_JSON — see
    drive_credentials(). File: DRIVE_FILE_ID (preferred) or TANKER_FILE_NAME
    (default "Tanker Billing.xlsm"). Returns a dict of source metadata.
    """
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    creds = drive_credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)

    file_id = os.environ.get("DRIVE_FILE_ID", "").strip()
    name = os.environ.get("TANKER_FILE_NAME", "Tanker Billing.xlsm").strip()
    if not file_id:
        q = f"name = '{name}' and trashed = false"
        resp = service.files().list(
            q=q, fields="files(id, name, modifiedTime)",
            includeItemsFromAllDrives=True, supportsAllDrives=True,
            orderBy="modifiedTime desc", pageSize=5).execute()
        files = resp.get("files", [])
        if not files:
            raise SystemExit(
                f"No Drive file named {name!r} is visible to these credentials. "
                "With GDRIVE_TOKEN, sign in as the account that owns the file; "
                "with GDRIVE_SA_JSON, share the file (or its folder) with the "
                "service account's email. Or set DRIVE_FILE_ID directly.")
        file_id = files[0]["id"]

    meta = service.files().get(
        fileId=file_id, fields="id, name, modifiedTime",
        supportsAllDrives=True).execute()

    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    import io
    with io.FileIO(dest_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _status, done = downloader.next_chunk()

    return {
        "drive_file_id": meta.get("id"),
        "file_name": meta.get("name"),
        "modified_time": meta.get("modifiedTime"),
    }


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--xlsm", help="path to a local Tanker Billing workbook")
    src.add_argument("--from-drive", action="store_true",
                     help="download the workbook from Google Drive first")
    ap.add_argument("--out", default=OUTPUT_PATH, help="output JSON path")
    args = ap.parse_args(argv)

    source = {"file_name": None, "drive_file_id": None, "modified_time": None}
    if args.from_drive:
        tmp = os.path.join(tempfile.gettempdir(), "tanker_billing.xlsm")
        source = download_from_drive(tmp)
        xlsm_path = tmp
    else:
        xlsm_path = args.xlsm
        source["file_name"] = os.path.basename(xlsm_path)

    if not os.path.exists(xlsm_path):
        raise SystemExit(f"workbook not found: {xlsm_path}")

    data = parse_workbook(xlsm_path)
    if not data["customers"]:
        raise SystemExit(
            "No customers parsed. Open the workbook in Excel, let the tables "
            "fill, save, and re-run (data_only needs cached formula values).")

    data["generated_at"] = dt.datetime.now(dt.timezone.utc).isoformat(
        timespec="seconds")
    data["source"] = source

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
        fh.write("\n")

    print(f"Wrote {len(data['customers'])} customers, "
          f"{len(data['rate_card'])} rate tiers, "
          f"{len(data['product_rates'])} product rates -> {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
