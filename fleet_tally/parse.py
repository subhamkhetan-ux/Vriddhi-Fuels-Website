"""Read the fleet-card settlement sheet: Date, Customer Name, Amount.

Tolerant to column order and to .xls/.xlsx; the header row is found by looking
for the Date / Customer / Amount labels. Each data row becomes a ``Row`` with a
parsed date, the customer name verbatim (it must match the Tally ledger exactly),
and a float amount; unparseable rows are returned with an ``error`` so the app
can show them instead of silently dropping them.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass


@dataclass
class Row:
    index: int
    date: dt.date | None
    customer: str
    amount: float
    error: str = ""


_DATE_FORMATS = [
    "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d", "%Y/%m/%d",
    "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d-%B-%Y", "%m/%d/%Y", "%m/%d/%y",
]


def _to_date(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_amount(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₹", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _match_col(headers: list[str], *needles: str) -> int | None:
    # Only treat a short, label-like cell as a header (so a long instruction
    # note that happens to mention "date"/"amount" is not mistaken for one).
    for i, h in enumerate(headers):
        hl = str(h or "").strip().lower()
        if len(hl) <= 30 and any(n in hl for n in needles):
            return i
    return None


def _find_header(grid: list[list]):
    """The header row is the first where Date, Customer and Amount labels land in
    three DISTINCT columns."""
    for i, r in enumerate(grid[:25]):
        ci_date = _match_col(r, "date")
        ci_cust = _match_col(r, "customer", "name", "party", "ledger")
        ci_amt = _match_col(r, "amount", "value", "amt")
        if None not in (ci_date, ci_cust, ci_amt) and len({ci_date, ci_cust, ci_amt}) == 3:
            return i, ci_date, ci_cust, ci_amt
    return None, None, None, None


def _rows_from_grid(grid: list[list]) -> list[Row]:
    header_i, ci_date, ci_cust, ci_amt = _find_header(grid)
    if header_i is None:
        return []

    out: list[Row] = []
    for n, r in enumerate(grid[header_i + 1:], start=1):
        def cell(ci):
            return r[ci] if ci < len(r) else None
        raw_cust = cell(ci_cust)
        cust = str(raw_cust).strip() if raw_cust is not None else ""
        date = _to_date(cell(ci_date))
        amt = _to_amount(cell(ci_amt))
        if not cust and date is None and amt is None:
            continue                       # blank spacer row
        err = ""
        if not cust:
            err = "missing customer name"
        elif date is None:
            err = "unreadable date"
        elif amt is None:
            err = "unreadable amount"
        elif amt == 0:
            err = "amount is zero"
        out.append(Row(index=n, date=date, customer=cust, amount=amt or 0.0, error=err))
    return out


def parse_excel(path: str) -> list[Row]:
    """Parse an .xlsx or .xls fleet-card sheet into ``Row`` objects."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        import xlrd
        book = xlrd.open_workbook(path)
        sh = book.sheet_by_index(0)
        grid = []
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                cell = sh.cell(i, j)
                if cell.ctype == 3:        # xlrd date
                    y, m, d, *_ = xlrd.xldate_as_tuple(cell.value, book.datemode)
                    row.append(dt.date(y, m, d))
                else:
                    row.append(cell.value)
            grid.append(row)
    return _rows_from_grid(grid)
