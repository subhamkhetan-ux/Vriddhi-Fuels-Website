#!/usr/bin/env python3
"""Export the canonical customer list to ``state/customers.json`` (maintenance).

The cloud runner can't open ``Master Ledger.xlsm`` from iCloud, so the canonical
names — the *computed* values of column F on "Master Paid" — are exported here on
the Mac and committed. Re-run this whenever new customers are added (spec §3).

    python3 export_customers.py            # default iCloud path
    python3 export_customers.py --xlsm /path/to/Master Ledger.xlsm
    python3 export_customers.py --column F --sheet "Master Paid"

Reads the workbook with cached formula values (``data_only``), so save the .xlsm
in Excel at least once after adding customers, so the spill formula's values are
cached in the file.
"""

from __future__ import annotations

import argparse
import os

from agent import state_store

DEFAULT_XLSM = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Vriddhi Fuels/Master Ledger.xlsm"
)


def export(xlsm_path: str, sheet: str, column: str) -> list[str]:
    import openpyxl  # local dep, only needed on the Mac

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet!r} not found; have: {wb.sheetnames}")
    ws = wb[sheet]
    col = openpyxl.utils.column_index_from_string(column)

    names: list[str] = []
    seen = set()
    for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
        val = row[0]
        if val is None:
            continue
        name = str(val).strip()
        # skip the header cell and blanks
        if not name or name.lower() in {"customer master list", "customer", "name"}:
            continue
        key = name.lower()
        if key not in seen:
            seen.add(key)
            names.append(name)
    return names


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Export Master Paid column F -> customers.json")
    ap.add_argument("--xlsm", default=DEFAULT_XLSM)
    ap.add_argument("--sheet", default="Master Paid")
    ap.add_argument("--column", default="F")
    args = ap.parse_args(argv)

    if not os.path.exists(args.xlsm):
        raise SystemExit(f"workbook not found: {args.xlsm}")

    names = export(args.xlsm, args.sheet, args.column)
    if not names:
        raise SystemExit(
            "No names found. Open the workbook in Excel, let column F fill, save, "
            "and re-run — data_only needs the cached formula values.")
    state_store.save_customers(names)
    print(f"Wrote {len(names)} customer name(s) to {state_store.CUSTOMERS_PATH}")
    print("Commit state/customers.json so the runner picks it up.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
