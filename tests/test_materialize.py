import json
import os

import openpyxl

import materialize
from agent import state_store
from agent.serial import EPOCH


def _seed(tmp_path, monkeypatch, rows, customers=None, aliases=None):
    """Point state_store at a temp dir and seed it."""
    state_dir = tmp_path / "state"
    state_dir.mkdir()
    monkeypatch.setattr(state_store, "QUEUE_PATH", str(state_dir / "queue.json"))
    monkeypatch.setattr(state_store, "SEEN_PATH", str(state_dir / "seen.json"))
    monkeypatch.setattr(state_store, "ALIASES_PATH", str(state_dir / "aliases.json"))
    monkeypatch.setattr(state_store, "CUSTOMERS_PATH", str(state_dir / "customers.json"))
    state_store.save_queue(rows)
    state_store.save_customers(customers or ["Sudarshan Minerals And Logistics"])
    state_store.save_aliases(aliases or {})
    return state_dir


MATCHED = {
    "entry_id": "abc123",
    "gmail_msg_id": "m1",
    "bank": "HDFC",
    "mode": "HDFC NEFT",
    "date_serial": 45831,
    "date_str": "15/06/2025",
    "amount": 125000,
    "raw_payer": "M/S SUDARSHAN MINERALS AND LOG",
    "customer": "Sudarshan Minerals And Logistics",
    "status": "matched",
}
REVIEW = {
    "entry_id": "def456",
    "gmail_msg_id": "m2",
    "bank": "ICICI",
    "mode": "ICICI UPI",
    "date_serial": 45832,
    "date_str": "16/06/2025",
    "amount": 5000,
    "raw_payer": "SOME UNKNOWN PAYER",
    "customer": None,
    "status": "review",
}


def test_materialize_writes_master_paid_shape(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch, [dict(MATCHED), dict(REVIEW)])
    rc = materialize.main(["--no-input", "--output-dir", str(tmp_path)])
    assert rc == 0

    files = [f for f in os.listdir(tmp_path) if f.startswith("PaymentEntries_")]
    assert len(files) == 1
    wb = openpyxl.load_workbook(tmp_path / files[0])
    ws = wb.active
    assert ws.title == "Master Paid"
    # header
    assert [ws.cell(1, c).value for c in range(1, 5)] == [
        "Date", "Customer", "Amount Paid", "Payment Mode"]
    # one data row (only the matched one), correct types.
    # Column A is a numeric serial with a date format; openpyxl reads such a
    # cell back as a datetime, but the stored value is the serial (like the
    # pay app's <v>45831</v>). Verify it round-trips to the same serial.
    a2 = ws.cell(2, 1)
    assert a2.is_date
    assert (a2.value.date() - EPOCH).days == 45831
    assert a2.number_format == "dd/mm/yyyy"
    assert ws.cell(2, 2).value == "Sudarshan Minerals And Logistics"
    assert ws.cell(2, 3).value == 125000           # plain number
    assert ws.cell(2, 4).value == "HDFC NEFT"
    assert ws.cell(3, 1).value is None             # review row not exported


def test_idempotent_rerun_produces_nothing(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch, [dict(MATCHED)])
    materialize.main(["--no-input", "--output-dir", str(tmp_path)])
    # matched row is now flagged materialized in the queue
    rows = state_store.load_queue()
    assert rows[0]["materialized"] is True

    # remove the produced file, re-run: no new file, nothing to export
    for f in os.listdir(tmp_path):
        if f.startswith("PaymentEntries_"):
            os.remove(tmp_path / f)
    materialize.main(["--no-input", "--output-dir", str(tmp_path)])
    assert not [f for f in os.listdir(tmp_path) if f.startswith("PaymentEntries_")]


def test_dry_run_changes_nothing(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch, [dict(MATCHED)])
    materialize.main(["--no-input", "--dry-run", "--output-dir", str(tmp_path)])
    assert not [f for f in os.listdir(tmp_path) if f.startswith("PaymentEntries_")]
    rows = state_store.load_queue()
    assert "materialized" not in rows[0]


def test_review_row_left_for_later(tmp_path, monkeypatch):
    _seed(tmp_path, monkeypatch, [dict(REVIEW)])
    materialize.main(["--no-input", "--output-dir", str(tmp_path)])
    rows = state_store.load_queue()
    assert rows[0]["status"] == "review"  # untouched in non-interactive mode
